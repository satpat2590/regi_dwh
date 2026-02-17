"""
This class will be used to take in a DataFrame and then format it as an Excel file seamlessly. 

You will be able to store the file in some output directory, and if not provided, it will be stored in your local directory in some folder 'output'.

You can also return a reference to the original data to perform some logical manipulation (i.e color coding, data transformation, aggregation, etc...)
"""

import sys, os
import pandas as pd
import openpyxl 
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
import pathlib


class ExcelFormatter:
    def __init__(self):
     # Create a reference to an empty Excel workbook
        self.wb = Workbook()
        self._table_names = set()
     

    def add_to_sheet(self, df: pd.DataFrame, sheet_name: str, transform_fn = None) -> None:
        """
        Pass in a dataframe object which will be added to a sheet in an Excel file object with a specific sheet name.\n

        :param df: The dataframe object which will be saved to the sheet
        :param sheet_name: The name of the sheet which will contain the data in df
        :param transform_fn: [IN PROGRESS] Add a lambda function which will take in a df as a paramater for pre-transformation prior to saving the data
        """

        ws = self.wb.active
     # Check if the default sheet is empty (only check first cell to avoid iterating millions of rows)
        is_empty = ws.max_row <= 1 and ws.cell(1, 1).value is None

        if is_empty:
         # Use the default sheet
            ws.title = sheet_name
        else:
         # Create a new sheet
            ws = self.wb.create_sheet(title=sheet_name)

     # Apply transformation to the dataframe
        if transform_fn:
            df = transform_fn(df)

     # Open up the sheet in the workbook
        if sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name] # Open up the sheet if it already exists 
        else:
            ws = self.wb.create_sheet(title=sheet_name)

     # Save the rows in the dataframe
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

     # Get metadata on rows and columns
        num_rows = ws.max_row
        num_cols = ws.max_column
        table_ref = f"A1:{get_column_letter(num_cols)}{num_rows}"

     # Create and style table with unique displayName
        display_name = "".join(sheet_name.split(" "))
        base_name = display_name
        counter = 2
        while display_name in self._table_names:
            display_name = f"{base_name}_{counter}"
            counter += 1
        self._table_names.add(display_name)
        table = Table(displayName=display_name, ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style 
        ws.add_table(table)

     # Format the column widths according to the size of the data in the rows
     # Sample up to 500 rows to avoid slow iteration on large DataFrames
        sample = df.head(500)
        for i, col in enumerate(df.columns, start=1):
            col_letter = get_column_letter(i)
            max_len = max(len(str(cell)) for cell in [col] + sample[col].astype(str).tolist())
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    def save(self, filename: str, location: str = None) -> None:
        """
        This method will take in a workbook object and then save it to a file location specified by 'location' \n
        If the 'location' parameter is empty, then it will store the output file to a folder called 'output' located in the relative directory. 

        :param filename: The name of the output Excel file which contains the resultant set from DB calls
        :param location: [OPTIONAL] The location in which the Excel file will be stored in the context of your directory structure. 
        """
     # Craft the full location path of the saved output file
        fpath = location if location else self.__create_output_dir()

     # Validate the filename to ensure it is .xlsx 
        path_split = filename.split('.')[1]
        if not path_split:
            print(f"There is no extension to the filename provided. \n\tPlease create a filename with the following example convention: example.xlsx")
            self.__reset_workbook()
            return None 
        elif path_split != 'xlsx':
            print(f"The file extension is '{path_split}' which is NOT of type 'xlsx'. Please correct it.")
            self.__reset_workbook()
            return None

     # Validate the save location to ensure it exists
        if not os.path.exists(fpath):
            print(f"The following file location does not exist: {location}. \n\tPlease enter a valid file location. ")
            self.__reset_workbook()
            return None

     # Save the file to the specified location if it exists
        spath = os.path.join(fpath, filename)
        self.wb.save(spath)
        self.__reset_workbook()

        print(f"Saved the output workbook succesfully to {spath}\n")

    def __create_output_dir(self) -> str:
        """
        Create a folder called 'output' in the relative directory.\n
        To be used in the 'save' method as a means to create the output directory seamlessly
        """

     # Get the base directory path 
        base_dir = str(pathlib.Path(__file__).parent)
     # Join the base_dir and the 'output' folder
        o_dir = base_dir + '\output'
     # Create a folder called 'output' in the base_dir
        if not os.path.exists(o_dir):
            try:
                os.mkdir(o_dir)
            except FileExistsError:
                print(f"Directory {o_dir} already exists.")
                pass # Continue with the logical flow
            except PermissionError:
                print(f"Permission denied: Unable to create '{o_dir}'")
                self.__reset_workbook()
                return None
            except Exception as e:
                print(f"An error occurred: {e}")
                self.__reset_workbook()
                return None

     # Return the output directory path 
        return os.path.realpath(o_dir)

    def __reset_workbook(self):
        print(f"Resetting the Excel workbook...")
        self.wb = Workbook()
        self._table_names = set()